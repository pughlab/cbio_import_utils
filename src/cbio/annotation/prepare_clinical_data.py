
from openpyxl import load_workbook
import argparse

def get_options():

    parser = argparse.ArgumentParser()
    parser.add_argument("-i", "--input", type=str, required=True,
                        help="input excel sheet inlcuding path")
    parser.add_argument("-o", "--output", type=str, required=True,
                        help="output file name including path")
    parser.add_argument("-s", "--sheet", type=str, required=True,
                        help="excel sheet name")
    parser.add_argument("-t", "--debug", action="store_true",
                        help="debug mode for testing")
    return parser.parse_args()

def main():
    args = get_options()
    print (args)
    if args.debug:
        print (args)

    # Load from the workbook
    wb = load_workbook(args.input)

    # Get sheet names
    if args.debug:
        print(wb.get_sheet_names())

    with open(args.output, 'w') as of:
        my_sheet = wb.get_sheet_by_name(args.sheet)

        for row in my_sheet.iter_rows(row_offset=0):
            parts = []
            for cell in row:
                if str(cell.value) == "None": continue
                parts.append(str(cell.value).strip())
            if len(parts) > 0:
                if args.debug:
                    print (parts)
                else:
                    of.write("{}\n".format("\t".join(parts)))

if __name__ == "__main__":
    main()